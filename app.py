from flask import Flask, render_template, request, redirect, url_for, session
from werkzeug.utils import secure_filename
from firebase_admin import credentials, firestore, initialize_app
from openpyxl import Workbook, load_workbook
import os
import json
import datetime

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Folder setup
UPLOAD_FOLDER = 'uploads'
EXCEL_PATH = 'data/responses.xlsx'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs('data', exist_ok=True)

# Firebase Initialization
cred = credentials.Certificate('firebase_key.json')  # Replace with your path
initialize_app(cred)
db = firestore.client()

# Routes
@app.route('/')
def vendor_form():
    return render_template('vendor_form.html')

@app.route('/submit_vendor', methods=['POST'])
def submit_vendor():
    session['vendor_data'] = {
        'Company Name': request.form.get('company_name'),
        'Vendor Name': request.form.get('vendor_name'),
        'Address': request.form.get('address'),
        'GSTIN': request.form.get('gstin'),
        'Contact Name': request.form.get('contact_name'),
        'Phone': request.form.get('phone'),
        'Email': request.form.get('email'),
        'Website': request.form.get('website'),
        'Payment Terms': request.form.get('payment_terms'),
        'Basis of Approval': request.form.get('basis_of_approval'),
        'Associated From': request.form.get('associated_from'),
        'Validity of Approval': request.form.get('validity_of_approval'),
        'Approved By': request.form.get('approved_by'),
        'Identification': request.form.get('identification'),
        'Feedback': request.form.get('feedback'),
        'Remarks': request.form.get('remarks'),
        'Enquired Part': request.form.get('enquired_part'),
        'Visited Date': request.form.get('visited_date'),
        'NDA Signed': request.form.get('nda_signed'),
        'Detailed Evaluation': request.form.get('detailed_evaluation'),
    }

    image = request.files.get('company_board')
    if image and image.filename:
        filename = secure_filename(image.filename)
        image_path = os.path.join(UPLOAD_FOLDER, filename)
        image.save(image_path)
        session['vendor_data']['Company Board Image'] = image_path
    else:
        session['vendor_data']['Company Board Image'] = ''

    session['machine_entries'] = []
    return redirect(url_for('machine_entry'))

@app.route('/machine_entry')
def machine_entry():
    machines = [
        'Vertical Turning Turret', 'Vertical Turning Ram', 'Turning Lathe', 'Turning CNC',
        'Conventional Milling', '3 Axis Vertical Machining Center', '3 Axis Horizontal Machining Center',
        '4 Axis Vertical Machining Center', '4 Axis Horizontal Machining Center', '4 Axis Turn Mill',
        '5 axis Milling', '5 Axis Mill Turn', '5 Axis Turn Mill', 'Surface Grinding',
        'Cylindrical Grinding', 'Spark Erosion Drill', 'Spark Electrical Discharge Machining',
        'Wire Electrical Discharge Machining'
    ]
    sizes = ['S1', 'S2', 'M1', 'M2', 'L1', 'L2']
    return render_template('machine_entry.html', machines=machines, sizes=sizes)

@app.route('/submit_machine', methods=['POST'])
def submit_machine():
    session['selected_machine'] = request.form.get('machine')
    session['selected_size'] = request.form.get('size')
    session['hour_rate'] = request.form.get('hour_rate')

    image = request.files.get('machine_image')
    if image and image.filename:
        filename = secure_filename(image.filename)
        image_path = os.path.join(UPLOAD_FOLDER, filename)
        image.save(image_path)
        session['machine_image'] = image_path
    else:
        session['machine_image'] = ''

    return redirect(url_for('specs_form'))

@app.route('/specs_form')
def specs_form():
    return render_template('specs_form.html', machine=session.get('selected_machine'), size=session.get('selected_size'))

@app.route('/submit_specs', methods=['POST'])
def submit_specs():
    data = {
        'Machine': session.get('selected_machine'),
        'Size': session.get('selected_size'),
        'Hour Rate': session.get('hour_rate'),
        'Machine Image': session.get('machine_image'),
        'Make': request.form.get('make'),
        'Model/Year': request.form.get('model_year'),
        'Type': request.form.get('type'),
        'Axis Configuration': request.form.get('axis_config'),
        'X-axis Travel': request.form.get('x_travel'),
        'Y-axis Travel': request.form.get('y_travel'),
        'Z’-axis Travel': request.form.get('z_travel'),
        'A’-axis Travel': request.form.get('a_travel'),
        'B’-axis Travel': request.form.get('b_travel'),
        'C’-axis Travel': request.form.get('c_travel'),
        'Max Part Size': request.form.get('max_part_size'),
        'Max Part Height': request.form.get('max_part_height'),
        'Spindle Taper': request.form.get('spindle_taper'),
        'Spindle Power': request.form.get('spindle_power'),
        'Spindle Torque': request.form.get('spindle_torque'),
        'Main Spindle Max RPM': request.form.get('main_spindle_rpm'),
        'Aux Spindle Max RPM': request.form.get('aux_spindle_rpm'),
        'Max Table Load': request.form.get('max_table_load'),
        'Thru Coolant Pressure': request.form.get('coolant_pressure'),
        'Pallet type': request.form.get('pallet_type'),
        'Positional Tolerance Standard': request.form.get('tolerance_std'),
        'Positional Accuracy X/Y/Z': request.form.get('accuracy_xyz'),
        'Positional Accuracy A/B/C': request.form.get('accuracy_abc'),
        'Positional Accuracy Table': request.form.get('accuracy_table'),
        'Angle Head': request.form.get('angle_head'),
        'Controller': request.form.get('controller'),
        'CAD Software': request.form.get('cad'),
        'CAM Software': request.form.get('cam')
    }

    if session.get('selected_machine') in [
        'Spark Erosion Drill', 'Spark Electrical Discharge Machining', 'Wire Electrical Discharge Machining'
    ]:
        data.update({
            'Wire Diameter': request.form.get('wire_dia'),
            'Taper Degree': request.form.get('taper_deg'),
            'Max Cutting Thickness': request.form.get('cutting_thickness'),
            'Surface Finish (Ra)': request.form.get('surface_finish'),
            'Electrode Diameter': request.form.get('electrode_dia'),
            'Spindle Stroke': request.form.get('spindle_stroke'),
            'Table Size': request.form.get('table_size'),
            'Sink Size': request.form.get('sink_size')
        })

    session['machine_entries'].append(data)
    session.modified = True

    if 'add_another' in request.form:
        return redirect(url_for('machine_entry'))
    else:
        return redirect(url_for('final_submit'))

@app.route('/final_submit')
def final_submit():
    vendor_data = session.get('vendor_data', {})
    machine_entries = session.get('machine_entries', [])

    if not machine_entries:
        return "No machines added", 400

    try:
        # FIRESTORE STORE
        vendor_ref = db.collection('vendors').document()
        vendor_ref.set({'vendor_data': vendor_data})

        for machine in machine_entries:
            machine_data = {
                key: machine[key]
                for key in ['Machine', 'Size', 'Hour Rate', 'Machine Image']
                if key in machine
            }
            specs_data = {
                key: val for key, val in machine.items()
                if key not in machine_data
            }
            vendor_ref.collection('machines').document().set({
                'machine_data': machine_data,
                'specs': specs_data
            })

        # EXCEL STORE
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            headers = ['Timestamp'] + list(vendor_data.keys()) + ['Machine', 'Size', 'Hour Rate']
            sample_keys = list(machine_entries[0].keys())
            for k in ['Machine', 'Size', 'Hour Rate']:
                if k in sample_keys:
                    sample_keys.remove(k)
            headers += sample_keys
            ws.append(headers)

        for machine in machine_entries:
            row = [datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            row += [vendor_data.get(k, '') for k in vendor_data]
            row += [machine.get('Machine'), machine.get('Size'), machine.get('Hour Rate')]
            for k in ws[1][len(row):]:  # Ensure columns match header
                row.append(machine.get(k.value, ''))
            ws.append(row)

        wb.save(EXCEL_PATH)

        session.clear()
        return "✅ Submission successful! Saved to Firestore and Excel."

    except Exception as e:
        return f"❌ Error: {str(e)}", 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)



